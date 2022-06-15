import os
 
import faker as faker
import openpyxl
from faker import Faker

separator = os.sep
excel_file_path = os.getcwd() + separator + "ExcelFiles" + separator + "DATA_DRIVEN_TESTING.xlsx"
print(excel_file_path)

workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active

faker = Faker()
print(faker.name())


for r in range(1,40):
    for c in range(1,5):
        if c == 1 :
            ids = faker.random_number(digits=3)
            sheet.cell(row=r, column=c).value = ids
        elif c == 2 :
            sheet.cell(row=r, column=c).value = faker.name()
        elif c == 3:
            sheet.cell(row=r, column=c).value = faker.name()
        else:
            salary = faker.random_number(digits=4)
            sheet.cell(row=r, column=c).value = salary


workbook.save(excel_file_path)



