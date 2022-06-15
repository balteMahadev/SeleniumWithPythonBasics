import os

import openpyxl

separator = os.sep
excel_file_path = os.getcwd() + separator + "ExcelFiles" + separator + "DATA_DRIVEN_TESTING.xlsx"
print(excel_file_path)

workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active  # workbook.get_sheet_by_name("Sheet1")
rows = sheet.max_row
cols = sheet.max_column
print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row =r, column=c).value, end="   ")
    print()